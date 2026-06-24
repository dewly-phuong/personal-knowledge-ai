import datetime
import os
import re

from pymongo import ReplaceOne

from app.services.mongodb_import_shared import (
    get_mtime,
    is_unchanged,
    open_db,
    row_hash,
    update_meta,
)


def import_xlsx_files_to_mongodb(dir_path: str = "raw/local"):
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required. Install it with: uv add openpyxl")

    xlsx_dir = os.path.join(dir_path, "csv")
    if not os.path.exists(xlsx_dir):
        print(f"XLSX directory {xlsx_dir} not found. Skipping XLSX import.")
        return {"status": "skipped", "message": "Directory not found."}

    db, meta_col = open_db()
    imported_sheets = skipped_files = skipped_sheets = 0
    for filename in sorted(os.listdir(xlsx_dir)):
        if not filename.lower().endswith((".xlsx", ".xlsm")):
            continue
        result = _import_xlsx_file(openpyxl, db, meta_col, xlsx_dir, filename)
        imported_sheets += result["imported_sheets"]
        skipped_files += result["skipped_files"]
        skipped_sheets += result["skipped_sheets"]
    return {
        "status": "success",
        "imported_sheets": imported_sheets,
        "skipped_files": skipped_files,
        "skipped_sheets": skipped_sheets,
    }


def _import_xlsx_file(openpyxl, db, meta_col, xlsx_dir: str, filename: str) -> dict:
    file_path = os.path.join(xlsx_dir, filename)
    meta_key = f"xlsx/{filename}"
    try:
        last_modified = get_mtime(file_path)
    except Exception as e:
        print(f"Failed to read file status for {filename}: {e}")
        return _result()
    if is_unchanged(meta_col, meta_key, last_modified):
        print(f"Skipping XLSX file (unchanged): {filename}")
        return _result(skipped_files=1)

    print(f"Importing XLSX file: {filename}")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"Failed to open {filename}: {e}")
        return _result()

    result = _result()
    file_sheets_imported = 0
    for sheet_name in wb.sheetnames:
        status = _import_sheet(db, wb[sheet_name], sheet_name)
        result["imported_sheets"] += int(status == "imported")
        result["skipped_sheets"] += int(status == "skipped")
        file_sheets_imported += int(status == "imported")
    update_meta(meta_col, meta_key, last_modified, sheets_imported=file_sheets_imported)
    return result


def _import_sheet(db, ws, sheet_name: str) -> str:
    all_rows = list(ws.iter_rows(values_only=True))
    parsed = _parse_sheet_rows(all_rows, sheet_name)
    if not parsed:
        return "skipped"
    rows, collection_name = parsed
    operations = []
    for row in rows:
        row["_row_key"] = row_hash(row)
        operations.append(ReplaceOne({"_row_key": row["_row_key"]}, row, upsert=True))
    try:
        db[collection_name].bulk_write(operations)
    except Exception as e:
        print(f"  Failed to write collection '{collection_name}': {e}")
        return "failed"
    return "imported"


def _parse_sheet_rows(all_rows: list, sheet_name: str):
    if len(all_rows) < 2:
        print(f"  Sheet '{sheet_name}': too few rows, skipping.")
        return None
    header_row_idx, headers = _detect_header(all_rows)
    if header_row_idx is None:
        print(f"  Sheet '{sheet_name}': could not detect header row, skipping.")
        return None
    rows = [_sheet_doc(headers, row) for row in all_rows[header_row_idx + 1 :]]
    rows = [row for row in rows if row]
    if not rows:
        print(f"  Sheet '{sheet_name}': all data rows are empty, skipping.")
        return None
    collection_name = _to_snake_case(sheet_name)
    print(
        f"  Sheet '{sheet_name}' -> collection '{collection_name}' ({len(rows)} rows)"
    )
    return rows, collection_name


def _detect_header(all_rows: list):
    for i, row in enumerate(all_rows):
        non_empty = [c for c in row if c is not None and str(c).strip()]
        if len(non_empty) >= max(1, len(row) * 0.5):
            headers = [
                _to_snake_case(str(c)) if c is not None else f"col_{j}"
                for j, c in enumerate(row)
            ]
            return i, headers
    return None, None


def _sheet_doc(headers: list[str], row: tuple) -> dict | None:
    cells = list(row) + [None] * max(0, len(headers) - len(row))
    doc, has_value = {}, False
    for header, value in zip(headers, cells):
        if isinstance(value, (datetime.datetime, datetime.date)):
            value = value.isoformat()
        doc[header] = value
        has_value = has_value or (value is not None and bool(str(value).strip()))
    return doc if has_value else None


def _to_snake_case(text: str) -> str:
    name = text.encode("ascii", errors="ignore").decode("ascii").strip()
    return re.sub(r"[^a-zA-Z0-9]+", "_", name).strip("_").lower()


def _result(imported_sheets: int = 0, skipped_files: int = 0, skipped_sheets: int = 0):
    return {
        "imported_sheets": imported_sheets,
        "skipped_files": skipped_files,
        "skipped_sheets": skipped_sheets,
    }
