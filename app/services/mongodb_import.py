import os
import re
import csv
import json
import hashlib
import datetime
from pymongo import MongoClient, ReplaceOne

def import_json_files_to_mongodb(dir_path: str = "raw/local"):
    """
    Scans dir_path for JSON files and imports/upserts them into MongoDB
    under the 'personal_knowledge_ai' database.
    Checks file modification dates to prevent duplicate imports.
    """
    mongo_uri = os.getenv("MONGO_URI", "mongodb://localhost:27017/")
    client = MongoClient(mongo_uri)
    db = client["personal_knowledge_ai"]
    metadata_col = db["_ingest_metadata"]

    if not os.path.exists(dir_path):
        print(f"Directory {dir_path} not found. Skipping JSON import.")
        return {"status": "skipped", "message": "Directory not found."}

    imported_count = 0
    skipped_count = 0

    for filename in os.listdir(dir_path):
        if not filename.endswith(".json"):
            continue

        file_path = os.path.join(dir_path, filename)
        collection_name = os.path.splitext(filename)[0]

        # Get last modified time
        try:
            mtime = os.path.getmtime(file_path)
            last_modified = datetime.datetime.fromtimestamp(mtime, datetime.timezone.utc).isoformat()
        except Exception as e:
            print(f"Failed to read file status for {filename}: {e}")
            continue

        # Check against metadata collection
        meta = metadata_col.find_one({"filepath": filename})
        if meta and meta.get("last_modified") == last_modified:
            skipped_count += 1
            continue

        # Load file content
        print(f"Importing JSON file to MongoDB: {filename} -> collection: {collection_name}")
        with open(file_path, "r", encoding="utf-8") as f:
            try:
                data = json.load(f)
            except Exception as e:
                print(f"Failed to parse JSON file {filename}: {e}")
                continue

        if not isinstance(data, list):
            data = [data]

        # Upsert elements using 'id' field
        operations = []
        for item in data:
            if isinstance(item, dict) and "id" in item:
                operations.append(
                    ReplaceOne({"id": item["id"]}, item, upsert=True)
                )
            else:
                # If no id, fallback to replacing the document if it matches completely, or insert
                # Using item as the filter
                operations.append(
                    ReplaceOne(item, item, upsert=True)
                )

        if operations:
            try:
                db[collection_name].bulk_write(operations)
            except Exception as e:
                print(f"Failed to bulk write documents for {filename}: {e}")
                continue

        # Update metadata
        metadata_col.update_one(
            {"filepath": filename},
            {"$set": {"last_modified": last_modified, "updated_at": datetime.datetime.now(datetime.timezone.utc).isoformat()}},
            upsert=True
        )
        imported_count += 1

    return {"status": "success", "imported": imported_count, "skipped": skipped_count}


def import_csv_files_to_mongodb(dir_path: str = "raw/local"):
    """
    Scans the 'csv/' subfolder inside dir_path for CSV files and imports/upserts
    them into MongoDB under the 'personal_knowledge_ai' database.
    Each CSV file becomes its own collection (filename without extension).
    Skips files that have not been modified since the last import.
    """
    mongo_uri = os.getenv("MONGO_URI", "mongodb://localhost:27017/")
    client = MongoClient(mongo_uri)
    db = client["personal_knowledge_ai"]
    metadata_col = db["_ingest_metadata"]

    csv_dir = os.path.join(dir_path, "csv")
    if not os.path.exists(csv_dir):
        print(f"CSV directory {csv_dir} not found. Skipping CSV import.")
        return {"status": "skipped", "message": "CSV directory not found."}

    imported_count = 0
    skipped_count = 0

    for filename in os.listdir(csv_dir):
        if not filename.endswith(".csv"):
            continue

        file_path = os.path.join(csv_dir, filename)
        collection_name = os.path.splitext(filename)[0]

        # Get last modified time
        try:
            mtime = os.path.getmtime(file_path)
            last_modified = datetime.datetime.fromtimestamp(mtime, datetime.timezone.utc).isoformat()
        except Exception as e:
            print(f"Failed to read file status for {filename}: {e}")
            continue

        # Check against metadata collection (keyed by filepath to avoid collision with JSON entries)
        meta_key = f"csv/{filename}"
        meta = metadata_col.find_one({"filepath": meta_key})
        if meta and meta.get("last_modified") == last_modified:
            skipped_count += 1
            print(f"Skipping CSV file (unchanged): {filename}")
            continue

        # Parse CSV
        print(f"Importing CSV file to MongoDB: {filename} -> collection: {collection_name}")
        rows = []
        try:
            with open(file_path, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # Strip whitespace from keys and values
                    rows.append({k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items()})
        except Exception as e:
            print(f"Failed to parse CSV file {filename}: {e}")
            continue

        if not rows:
            print(f"  No rows found in {filename}. Skipping.")
            continue

        # Build a stable, unique key per row by hashing all field values.
        # This is correct for ANY CSV layout, including time-series data where
        # a column like 'employee_id' is repeated across multiple rows.
        operations = []
        for row in rows:
            row_hash = hashlib.sha256(
                json.dumps(row, sort_keys=True, ensure_ascii=False).encode("utf-8")
            ).hexdigest()
            row["_row_key"] = row_hash
            operations.append(ReplaceOne({"_row_key": row_hash}, row, upsert=True))

        if operations:
            try:
                db[collection_name].bulk_write(operations)
            except Exception as e:
                print(f"Failed to bulk write documents for {filename}: {e}")
                continue

        # Update metadata
        metadata_col.update_one(
            {"filepath": meta_key},
            {"$set": {
                "last_modified": last_modified,
                "updated_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
                "row_count": len(rows),
            }},
            upsert=True
        )
        imported_count += 1

    return {"status": "success", "imported": imported_count, "skipped": skipped_count}


def _to_snake_case(text: str) -> str:
    """Converts a header string to snake_case (ASCII-safe, spaces/hyphens to underscores)."""
    name = text.encode("ascii", errors="ignore").decode("ascii").strip()
    return re.sub(r"[^a-zA-Z0-9]+", "_", name).strip("_").lower()


def _xlsx_sheet_to_collection_name(sheet_name: str) -> str:
    """Converts an Excel sheet name to a valid MongoDB collection name (snake_case, ASCII-only)."""
    return _to_snake_case(sheet_name)


def import_xlsx_files_to_mongodb(dir_path: str = "raw/local"):
    """
    Scans the 'csv/' subfolder inside dir_path for .xlsx/.xlsm files.
    Each sheet becomes its own MongoDB collection under 'personal_knowledge_ai'.
    Sheets with no detectable header or no data rows are skipped (e.g. summary/title sheets).
    Uses hash-based dedup and _ingest_metadata change tracking identical to the CSV importer.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required. Install it with: uv add openpyxl")

    mongo_uri = os.getenv("MONGO_URI", "mongodb://localhost:27017/")
    client = MongoClient(mongo_uri)
    db = client["personal_knowledge_ai"]
    metadata_col = db["_ingest_metadata"]

    xlsx_dir = os.path.join(dir_path, "csv")
    if not os.path.exists(xlsx_dir):
        print(f"XLSX directory {xlsx_dir} not found. Skipping XLSX import.")
        return {"status": "skipped", "message": "Directory not found."}

    imported_sheets = 0
    skipped_files = 0
    skipped_sheets = 0

    for filename in sorted(os.listdir(xlsx_dir)):
        if not filename.lower().endswith((".xlsx", ".xlsm")):
            continue

        file_path = os.path.join(xlsx_dir, filename)

        try:
            mtime = os.path.getmtime(file_path)
            last_modified = datetime.datetime.fromtimestamp(mtime, datetime.timezone.utc).isoformat()
        except Exception as e:
            print(f"Failed to read file status for {filename}: {e}")
            continue

        meta_key = f"xlsx/{filename}"
        meta = metadata_col.find_one({"filepath": meta_key})
        if meta and meta.get("last_modified") == last_modified:
            skipped_files += 1
            print(f"Skipping XLSX file (unchanged): {filename}")
            continue

        print(f"Importing XLSX file: {filename}")
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            print(f"Failed to open {filename}: {e}")
            continue

        file_sheets_imported = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))

            if len(all_rows) < 2:
                print(f"  Sheet '{sheet_name}': too few rows, skipping.")
                skipped_sheets += 1
                continue

            # Find the header row: first row where >=50% of cells are non-empty
            header_row_idx = None
            headers = None
            for i, row in enumerate(all_rows):
                non_empty = [c for c in row if c is not None and str(c).strip()]
                if len(non_empty) >= max(1, len(row) * 0.5):
                    header_row_idx = i
                    headers = [
                        _to_snake_case(str(c)) if c is not None else f"col_{j}"
                        for j, c in enumerate(row)
                    ]
                    break

            if header_row_idx is None:
                print(f"  Sheet '{sheet_name}': could not detect header row, skipping.")
                skipped_sheets += 1
                continue

            data_rows = all_rows[header_row_idx + 1:]
            if not data_rows:
                print(f"  Sheet '{sheet_name}': no data rows, skipping.")
                skipped_sheets += 1
                continue

            rows = []
            for row in data_rows:
                # Pad row to header length if needed
                cells = list(row) + [None] * max(0, len(headers) - len(row))
                doc = {}
                has_value = False
                for h, v in zip(headers, cells):
                    if isinstance(v, (datetime.datetime, datetime.date)):
                        v = v.isoformat()
                    doc[h] = v
                    if v is not None and str(v).strip():
                        has_value = True
                if has_value:
                    rows.append(doc)

            if not rows:
                print(f"  Sheet '{sheet_name}': all data rows are empty, skipping.")
                skipped_sheets += 1
                continue

            collection_name = _xlsx_sheet_to_collection_name(sheet_name)
            print(f"  Sheet '{sheet_name}' -> collection '{collection_name}' ({len(rows)} rows)")

            operations = []
            for row in rows:
                row_hash = hashlib.sha256(
                    json.dumps(row, sort_keys=True, ensure_ascii=False, default=str).encode("utf-8")
                ).hexdigest()
                row["_row_key"] = row_hash
                operations.append(ReplaceOne({"_row_key": row_hash}, row, upsert=True))

            try:
                db[collection_name].bulk_write(operations)
                file_sheets_imported += 1
                imported_sheets += 1
            except Exception as e:
                print(f"  Failed to write collection '{collection_name}': {e}")

        metadata_col.update_one(
            {"filepath": meta_key},
            {"$set": {
                "last_modified": last_modified,
                "updated_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
                "sheets_imported": file_sheets_imported,
            }},
            upsert=True
        )

    return {
        "status": "success",
        "imported_sheets": imported_sheets,
        "skipped_files": skipped_files,
        "skipped_sheets": skipped_sheets,
    }
